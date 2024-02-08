import win32com.client
import pywintypes
import pandas
import numpy
import datetime
import shutil
import re
import time
import os
import psutil
import numpy as np
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

cgl = get_column_letter
cifs = column_index_from_string
xlPasteFormats                 = -4122
xlPasteSpecialOperationNone    = -4142
xlDown = -4121 # https://learn.microsoft.com/en-us/office/vba/api/excel.xldirection

def countdown_timer(t):
  t = int(t)
  while t:
    mins, secs = divmod(t, 60)
    timer = '{:02d}:{:02d}'.format(mins, secs)
    print('Sleeping for', timer, end='\r')
    time.sleep(1)
    t -= 1

def set_value(ws, cell, value, activate_first = True):
  if activate_first:
    ws.Activate()
  to_excel = [[value]]
  height = len(to_excel)
  xy = coordinate_from_string(cell) # returns ('A',4)
  sx = column_index_from_string(xy[0]) # col
  sy = xy[1] # row
  write_range = ws.Range(ws.Cells(sy, sx), ws.Cells(sy + height - 1, sx))
  write_range.Value = to_excel
  
def get_value(ws, cell):
  return ws.Range(cell).Value

def get_formula(ws, cell):
  return ws.Range(cell).Formula

def get_range(ws, range, kind='values'):
  if kind.lower() == 'values':
    return list(list(ii) for ii in ws.Range(range).Value)
  elif kind.lower() == 'formulas':
    return list(list(ii) for ii in ws.Range(range).Formula)

def df_from_range(ws, range, kind = 'values', header=True):
  df = pandas.DataFrame(get_range(ws, range, kind))
  if header:
    df.columns = df.iloc[0]
    df = df[1:]
  return df
 
def set_range(ws, trange='auto~A1', to_excel=[['Howdy', 'There']], kind='values', activate_first = True):
  if activate_first:
    ws.Activate()
  if type(to_excel) == pandas.core.frame.DataFrame:
    te = to_excel.values.tolist()
  else:
    te = to_excel
  if 'auto' in trange:
    start_cell = trange.split('~')[-1]
    rng = calculate_destination_range_from_dataframe(start_cell, te)
  else:
    rng = trange
  if kind.lower() == 'values':
    ws.Range(rng).Value = te
  elif kind.lower() == 'formulas':
    ws.Range(rng).Formula = te
    
def calculate_destination_range_from_dataframe(start_cell, df_or_list, drop_first_n_rows=0):
  if type(df_or_list) == list:
    wide = len(df_or_list[0])
  else: # assume pandas dataframe
    wide = len(df_or_list.columns)
  strt_col, strt_row = coordinate_from_string(start_cell)
  end_col_num = cifs(strt_col) + wide - 1
  end_row = strt_row + len(df_or_list) - 1
  end_cell = get_column_letter(end_col_num) + str(end_row)
  
  modded_strt_row = strt_row + drop_first_n_rows
  modded_start_cell = strt_col + str(modded_strt_row)
  
  total_range = f'{modded_start_cell}:{end_cell}'
  return total_range

def paint_format(ws1, ws2, source_range, target_range):
  if ws2 == None:
    ws2 = ws1
  cells1 = ws1.Range(source_range)
  cells2 = ws2.Range(target_range)
  cells1.Copy()
  cells2.PasteSpecial(xlPasteFormats, xlPasteSpecialOperationNone)
  
def openWorkbook(xlapp, xlfile):
  print(f'Opening workbook "{os.path.basename(xlfile)}"')
  try: # try accessing directly first in case already open
    xlwb = xlapp.Workbooks(os.path.basename(xlfile))
    # print('a')
  except Exception as e:
    try: # open from disk if not already open
      xlwb = xlapp.Workbooks.Open(xlfile)
      # print('b')
    except Exception as e:
      print('c')
      print(e)
      xlwb = None
  return(xlwb)
  
def add_cols(start_col, num_cols):
  st = column_index_from_string(start_col)
  en = st + num_cols
  return get_column_letter(en)

def get_last_cell_in_column(xlapp, ws, start_cell):
  ws.Activate()
  time.sleep(0.1)
  ws.Range(start_cell).select
  time.sleep(0.1)
  ws.Range(start_cell).End(xlDown).select
  addr = xlapp.ActiveCell.Address.replace("$", "")
  return (addr, coordinate_from_string(addr))
  
def kill_excel():
  for proc in psutil.process_iter():
    if 'EXCEL.EXE' in proc.name():
      print('Killing', proc.name(), 'at PID', proc.pid)
      proc.terminate()
      break

def get_excel(kill_excel_first=False):
  if kill_excel_first:
    kill_excel()
  try:
    excel = win32com.client.GetActiveObject('Excel.Application')
  except Exception as e:
    excel = win32com.client.Dispatch('Excel.Application')
    excel.Visible = True
  return excel

def rgb_to_excel_color(rgb_tuple):
  red, green, blue = rgb_tuple
  bgr = (blue, green, red)  # Excel uses BGR order
  hex_color = '%02x%02x%02x' % bgr  # Convert to hex string
  return int(hex_color, 16)  # Convert hex string to integer
